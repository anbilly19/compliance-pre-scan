import io
from openpyxl import load_workbook
from .base import BaseExtractor, ExtractionResult


class XlsxExtractor(BaseExtractor):
    """Extract text from XLSX files."""

    def extract(self, data: bytes, filename: str) -> ExtractionResult:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

        cell_values: list[str] = []
        sheet_count = len(wb.sheetnames)

        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        cell_values.append(str(cell))

        wb.close()

        # openpyxl exposes macro detection via workbook.vba_archive
        has_macros = False
        try:
            wb2 = load_workbook(io.BytesIO(data), read_only=False)
            has_macros = wb2.vba_archive is not None
            wb2.close()
        except Exception:
            pass

        return ExtractionResult(
            text="\n".join(cell_values),
            page_count=sheet_count,
            raw_byte_size=len(data),
            has_macros=has_macros,
        )
